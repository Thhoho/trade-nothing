#!/usr/bin/env python3
"""
Trade Nothing v0.9 — DCF Financial Model Excel Builder

Generates professional, institutional-grade, formula-driven Excel models.
Adheres strictly to "Formulas Over Hardcodes" and 5×5 WACC/g Sensitivity Analysis.

Usage:
  python3 excel_model_builder.py --code 300118 --name "Target Co" --price 15.0 \
      --shares 1140 --net-debt 4500
  python3 excel_model_builder.py --code AAPL --name "Apple" --price 195 \
      --shares 15400 --net-debt -60000 --out ~/my_model.xlsx
"""

import os
import sys
import re
import argparse
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils import get_output_dir

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
except ImportError:
    print("[ERROR] openpyxl is required. Install: pip install openpyxl>=3.1.0", file=sys.stderr)
    sys.exit(1)


def generate_slug(text):
    words = re.findall(r"[\u4e00-\u9fa5\w]+", text.lower())
    return "_".join(words)[:15]


def build_dcf_model(filepath, code, name, current_price, shares, net_debt,
                    wacc, terminal_g, tax_rate, hist_rev, growth_rates, ebit_margins):
    os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DCF Valuation"
    ws.views.sheetView[0].showGridLines = True

    # Design tokens (Investment Banking theme)
    FONT_FAMILY = "Segoe UI"
    font_title = Font(name=FONT_FAMILY, size=15, bold=True, color="FFFFFF")
    font_section = Font(name=FONT_FAMILY, size=11, bold=True, color="1F4E79")
    font_header = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
    font_bold = Font(name=FONT_FAMILY, size=10, bold=True, color="000000")
    font_regular = Font(name=FONT_FAMILY, size=10, bold=False, color="000000")
    font_input = Font(name=FONT_FAMILY, size=10, color="0070C0", bold=False)
    font_input_bold = Font(name=FONT_FAMILY, size=10, color="0070C0", bold=True)

    fill_dark_navy = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_soft_blue = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_center_base = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    border_light = Border(
        left=Side(style="thin", color="E0E0E0"), right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"), bottom=Side(style="thin", color="E0E0E0")
    )
    border_header = Border(
        top=Side(style="medium", color="1F4E79"), bottom=Side(style="medium", color="1F4E79")
    )

    # Title Block
    ws.merge_cells("A1:H2")
    title_cell = ws["A1"]
    title_cell.value = f"TRADE NOTHING v0.9 — DCF VALUATION MODEL: {name} ({code})"
    title_cell.font = font_title
    title_cell.fill = fill_dark_navy
    title_cell.alignment = align_center
    for row in range(1, 3):
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = fill_dark_navy

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20

    # Metadata section (A4:C)
    ws.merge_cells("A4:C4")
    ws["A4"].value = "VALUATION METADATA & SUMMARY"
    ws["A4"].font = font_section
    ws["A4"].fill = fill_soft_blue
    ws["A4"].alignment = align_left

    metadata_rows = [
        ("Valuation Date", "=TODAY()", "YYYY-MM-DD", False),
        ("Current Share Price", current_price, "¥#,##0.00", True),
        ("Shares Outstanding (M)", shares, "#,##0.0", True),
        ("Net Debt (M)", net_debt, "¥#,##0.0", True),
        ("Implied Enterprise Value (M)", "=B34", "¥#,##0.0", False),
        ("Implied Equity Value (M)", "=B36", "¥#,##0.0", False),
        ("Implied Value / Share", "=B38", "¥#,##0.00", False),
        ("Premium / (Discount)", "=B40", "0.0%", False),
        ("Kelly Sizing Suggestion", 0.125, "0.0%", True),
    ]

    for i, (label, val, num_fmt, is_input) in enumerate(metadata_rows):
        row_idx = 5 + i
        ws.cell(row=row_idx, column=1, value=label).font = font_regular
        ws.cell(row=row_idx, column=1).alignment = align_left
        ws.cell(row=row_idx, column=1).border = border_light
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        ws.cell(row=row_idx, column=2).border = border_light
        val_cell = ws.cell(row=row_idx, column=3, value=val)
        val_cell.font = font_input if is_input else font_bold
        val_cell.alignment = align_right
        val_cell.number_format = num_fmt
        val_cell.border = border_light

    # Core Assumptions (E4:H)
    ws.merge_cells("E4:H4")
    ws["E4"].value = "CORE VALUATION ASSUMPTIONS"
    ws["E4"].font = font_section
    ws["E4"].fill = fill_soft_blue
    ws["E4"].alignment = align_left

    assumptions = [
        ("Base WACC", wacc, "0.0%", True),
        ("Terminal Growth Rate (g)", terminal_g, "0.0%", True),
        ("Effective Tax Rate", tax_rate, "0.0%", True),
        ("Historical Revenue (Y0) (M)", hist_rev, "¥#,##0.0", True),
    ]

    for i, (label, val, num_fmt, is_input) in enumerate(assumptions):
        row_idx = 5 + i
        ws.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=6)
        ws.cell(row=row_idx, column=5, value=label).font = font_regular
        ws.cell(row=row_idx, column=5).alignment = align_left
        ws.cell(row=row_idx, column=5).border = border_light
        ws.cell(row=row_idx, column=6).border = border_light
        ws.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=8)
        val_cell = ws.cell(row=row_idx, column=7, value=val)
        val_cell.font = font_input if is_input else font_bold
        val_cell.alignment = align_right
        val_cell.number_format = num_fmt
        val_cell.border = border_light
        ws.cell(row=row_idx, column=8).border = border_light

    for idx, r_growth in enumerate(growth_rates):
        row_idx = 9 + idx
        ws.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=6)
        ws.cell(row=row_idx, column=5, value=f"Year {idx+1} Revenue Growth").font = font_regular
        ws.cell(row=row_idx, column=5).alignment = align_left
        ws.cell(row=row_idx, column=5).border = border_light
        ws.cell(row=row_idx, column=6).border = border_light
        ws.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=8)
        val_cell = ws.cell(row=row_idx, column=7, value=r_growth)
        val_cell.font = font_input
        val_cell.alignment = align_right
        val_cell.number_format = "0.0%"
        val_cell.border = border_light
        ws.cell(row=row_idx, column=8).border = border_light

    # Financial Projection Schedule (Row 14)
    ws.merge_cells("A14:H14")
    ws["A14"].value = "FINANCIAL PROJECTION SCHEDULE (5-YEAR FORECAST)"
    ws["A14"].font = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
    ws["A14"].fill = fill_dark_navy
    ws["A14"].alignment = align_left
    for col in range(1, 9):
        ws.cell(row=14, column=col).fill = fill_dark_navy

    headers = ["Metric", "Driver / Formula", "Year 0", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    for col_idx, h in enumerate(headers):
        cell = ws.cell(row=15, column=col_idx + 1, value=h)
        cell.font = font_header
        cell.fill = fill_dark_navy
        cell.alignment = align_center
        cell.border = border_header

    # Revenue Growth (Row 16)
    ws.cell(row=16, column=1, value="Revenue Growth Rate").font = font_bold
    ws.cell(row=16, column=2, value="Assumption Inputs").font = font_regular
    ws.cell(row=16, column=3, value="N/A").font = font_regular
    ws.cell(row=16, column=3).alignment = align_right
    for y in range(1, 6):
        cell = ws.cell(row=16, column=3 + y, value=f"=G{8 + y}")
        cell.font = font_bold
        cell.alignment = align_right
        cell.number_format = "0.0%"

    # Revenue (Row 17)
    ws.cell(row=17, column=1, value="Revenue (M)").font = font_bold
    ws.cell(row=17, column=2, value="Prior × (1 + Growth)").font = font_regular
    ws.cell(row=17, column=3, value="=G8").font = font_bold
    ws.cell(row=17, column=3).alignment = align_right
    ws.cell(row=17, column=3).number_format = "¥#,##0.0"
    for y in range(1, 6):
        col_l = get_column_letter(3 + y)
        prev_l = get_column_letter(2 + y)
        cell = ws.cell(row=17, column=3 + y, value=f"={prev_l}17*(1+{col_l}16)")
        cell.font = font_bold
        cell.alignment = align_right
        cell.number_format = "¥#,##0.0"

    # EBIT Margin (Row 18), EBIT (19), Taxes (20), NOPAT (21), D&A (22), CapEx (23), NWC (24), FCF (25)
    row_defs = [
        (18, "EBIT Margin", "Targets", ebit_margins, "0.0%", font_input),
        (19, "EBIT (M)", "Revenue × Margin", None, "¥#,##0.0", font_bold),
        (20, "Taxes (M)", "EBIT × Tax Rate", None, "¥#,##0.0", font_regular),
        (21, "NOPAT (M)", "EBIT - Taxes", None, "¥#,##0.0", font_bold),
        (22, "D&A (M)", "4.0% of Revenue", None, "¥#,##0.0", font_regular),
        (23, "CapEx (M)", "4.5% of Revenue", None, "¥#,##0.0", font_regular),
        (24, "NWC Change (M)", "1.0% of Revenue", None, "¥#,##0.0", font_regular),
        (25, "Free Cash Flow (M)", "NOPAT+D&A+CapEx+NWC", None, "¥#,##0.0", font_bold),
    ]

    for row, label, driver, values, fmt, fnt in row_defs:
        ws.cell(row=row, column=1, value=label).font = fnt
        ws.cell(row=row, column=2, value=driver).font = font_regular

    # EBIT Margin values
    ws.cell(row=18, column=3, value=0.08).font = font_bold
    ws.cell(row=18, column=3).alignment = align_right
    ws.cell(row=18, column=3).number_format = "0.0%"
    for y in range(1, 6):
        cell = ws.cell(row=18, column=3 + y, value=ebit_margins[y-1])
        cell.font = font_input
        cell.alignment = align_right
        cell.number_format = "0.0%"

    # Formula rows
    for y in range(0, 6):
        cl = get_column_letter(3 + y)
        for row, formula in [
            (19, f"={cl}17*{cl}18"),
            (20, f"={cl}19*$G$7"),
            (21, f"={cl}19-{cl}20"),
            (22, f"={cl}17*0.04"),
            (23, f"=-({cl}17*0.045)"),
            (24, f"=-({cl}17*0.01)"),
            (25, f"={cl}21+{cl}22+{cl}23+{cl}24"),
        ]:
            cell = ws.cell(row=row, column=3 + y, value=formula)
            cell.alignment = align_right
            cell.number_format = "¥#,##0.0"

    # Discount rows (26-28)
    ws.cell(row=26, column=1, value="Discount Period (Years)").font = font_regular
    ws.cell(row=26, column=2, value="Forecast Year Index").font = font_regular
    ws.cell(row=26, column=3, value="N/A").font = font_regular
    ws.cell(row=26, column=3).alignment = align_right
    for y in range(1, 6):
        ws.cell(row=26, column=3 + y, value=y).font = font_regular
        ws.cell(row=26, column=3 + y).alignment = align_right

    ws.cell(row=27, column=1, value="Discount Factor").font = font_regular
    ws.cell(row=27, column=2, value="1/(1+WACC)^n").font = font_regular
    ws.cell(row=27, column=3, value="N/A").font = font_regular
    ws.cell(row=27, column=3).alignment = align_right
    for y in range(1, 6):
        cl = get_column_letter(3 + y)
        cell = ws.cell(row=27, column=3 + y, value=f"=1/((1+$G$5)^{cl}26)")
        cell.font = font_regular
        cell.alignment = align_right
        cell.number_format = "0.0000"

    ws.cell(row=28, column=1, value="PV of FCF (M)").font = font_bold
    ws.cell(row=28, column=2, value="FCF × Discount Factor").font = font_regular
    ws.cell(row=28, column=3, value="N/A").font = font_regular
    ws.cell(row=28, column=3).alignment = align_right
    for y in range(1, 6):
        cl = get_column_letter(3 + y)
        cell = ws.cell(row=28, column=3 + y, value=f"={cl}25*{cl}27")
        cell.font = font_bold
        cell.alignment = align_right
        cell.number_format = "¥#,##0.0"

    # Apply borders
    for row in range(15, 29):
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = border_light

    # DCF Outputs (Row 30)
    ws.merge_cells("A30:C30")
    ws["A30"].value = "DCF VALUATION OUTPUTS"
    ws["A30"].font = font_section
    ws["A30"].fill = fill_soft_blue
    ws["A30"].alignment = align_left

    outputs = [
        ("Cumulative PV of FCF", "=SUM(D28:H28)", "¥#,##0.0"),
        ("Terminal Value (TV)", "=H25*(1+$G$6)/($G$5-$G$6)", "¥#,##0.0"),
        ("PV of Terminal Value", "=B32*H27", "¥#,##0.0"),
        ("Implied Enterprise Value", "=B31+B33", "¥#,##0.0"),
        ("Less: Net Debt (M)", "=C8", "¥#,##0.0"),
        ("Implied Equity Value", "=B34-B35", "¥#,##0.0"),
        ("Shares Outstanding (M)", "=C7", "#,##0.0"),
        ("Implied Value / Share", "=B36/B37", "¥#,##0.00"),
        ("Current Share Price", "=C6", "¥#,##0.00"),
        ("Premium / (Discount)", "=(B38-B39)/B39", "0.0%"),
    ]

    for i, (label, val, num_fmt) in enumerate(outputs):
        row_idx = 31 + i
        ws.cell(row=row_idx, column=1, value=label).font = font_bold if "Value" in label else font_regular
        ws.cell(row=row_idx, column=1).alignment = align_left
        ws.cell(row=row_idx, column=1).border = border_light
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)
        val_cell = ws.cell(row=row_idx, column=2, value=val)
        val_cell.font = font_bold
        val_cell.alignment = align_right
        val_cell.number_format = num_fmt
        val_cell.border = border_light
        ws.cell(row=row_idx, column=3).border = border_light

    # Sensitivity Table (E30:J)
    ws.merge_cells("E30:J30")
    ws["E30"].value = "SENSITIVITY: WACC vs TERMINAL GROWTH"
    ws["E30"].font = font_section
    ws["E30"].fill = fill_soft_blue
    ws["E30"].alignment = align_left

    ws.cell(row=31, column=5, value="WACC / g").font = font_header
    ws.cell(row=31, column=5).fill = fill_dark_navy
    ws.cell(row=31, column=5).alignment = align_center
    ws.cell(row=31, column=5).border = border_light

    g_formulas = [("F31", "=$G$6-0.01"), ("G31", "=$G$6-0.005"), ("H31", "=$G$6"),
                  ("I31", "=$G$6+0.005"), ("J31", "=$G$6+0.01")]
    for cell_id, formula in g_formulas:
        cell = ws[cell_id]
        cell.value = formula
        cell.font = font_header
        cell.fill = fill_dark_navy
        cell.alignment = align_center
        cell.number_format = "0.0%"
        cell.border = border_light

    wacc_formulas = [("E32", "=$G$5-0.01"), ("E33", "=$G$5-0.005"), ("E34", "=$G$5"),
                     ("E35", "=$G$5+0.005"), ("E36", "=$G$5+0.01")]
    for cell_id, formula in wacc_formulas:
        cell = ws[cell_id]
        cell.value = formula
        cell.font = font_header
        cell.fill = fill_dark_navy
        cell.alignment = align_center
        cell.number_format = "0.0%"
        cell.border = border_light

    # 5×5 sensitivity grid
    for row_num in range(32, 37):
        wacc_ref = f"$E{row_num}"
        for col_num in range(6, 11):
            col_letter = get_column_letter(col_num)
            g_ref = f"{col_letter}$31"
            formula = (
                f"=( ( ($D$25/(1+{wacc_ref})^1 + $E$25/(1+{wacc_ref})^2 + $F$25/(1+{wacc_ref})^3 "
                f"+ $G$25/(1+{wacc_ref})^4 + $H$25/(1+{wacc_ref})^5) "
                f"+ ( ($H$25*(1+{g_ref})/({wacc_ref}-{g_ref})) / (1+{wacc_ref})^5 ) ) - $C$8) / $C$7"
            )
            cell = ws.cell(row=row_num, column=col_num, value=formula)
            cell.alignment = align_right
            cell.number_format = "¥#,##0.00"
            cell.border = border_light
            if row_num == 34 and col_num == 8:
                cell.font = font_input_bold
                cell.fill = fill_center_base
                cell.comment = Comment("Center = Base Case implied value.", "Trade Nothing v0.9")
            else:
                cell.font = font_regular

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22
    for col_idx in range(3, 11):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    wb.save(filepath)
    print(f"[SUCCESS] DCF Excel Model saved: {filepath}")


def main():
    parser = argparse.ArgumentParser(description="Trade Nothing Excel DCF Builder v0.9")
    parser.add_argument("--code", default="300118", help="Stock code")
    parser.add_argument("--name", default="Target Co", help="Stock name")
    parser.add_argument("--price", type=float, default=15.0, help="Current price")
    parser.add_argument("--shares", type=float, default=1140.0, help="Shares outstanding (M)")
    parser.add_argument("--net-debt", type=float, default=4500.0, help="Net Debt (M)")
    parser.add_argument("--wacc", type=float, default=0.09, help="WACC")
    parser.add_argument("--g", type=float, default=0.02, help="Terminal growth rate")
    parser.add_argument("--tax", type=float, default=0.15, help="Tax rate")
    parser.add_argument("--rev", type=float, default=35300.0, help="Year 0 Revenue (M)")
    parser.add_argument("--out", default="", help="Output filepath")
    args = parser.parse_args()

    if not args.out:
        out_dir = get_output_dir()
        os.makedirs(out_dir, exist_ok=True)
        args.out = os.path.join(out_dir, f"{args.code}_{generate_slug(args.name)}_dcf_model.xlsx")

    growth_rates = [0.12, 0.10, 0.08, 0.06, 0.05]
    ebit_margins = [0.09, 0.095, 0.10, 0.10, 0.10]

    build_dcf_model(
        filepath=args.out, code=args.code, name=args.name,
        current_price=args.price, shares=args.shares, net_debt=args.net_debt,
        wacc=args.wacc, terminal_g=args.g, tax_rate=args.tax,
        hist_rev=args.rev, growth_rates=growth_rates, ebit_margins=ebit_margins
    )


if __name__ == "__main__":
    main()
